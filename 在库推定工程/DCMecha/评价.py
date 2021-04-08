import numpy as np
import openpyxl
import progressbar
from MysqlApl import DB
from copy import deepcopy
import os
import 获取数据 as Get
path_excel = "C:/Users/liang-jiashun/Desktop/Excel模板/九宫格评价.xlsx"
path = "C:/Users/liang-jiashun/Desktop/库存感知测试结果汇总_最终"


def main():
    excel(size_min=3, size_max=3)


class Score:
    __size_max = 4
    __size_min = 3

    def __init__(self, size_max=4, size_min=3):
        self.__size_max = size_max
        self.__size_min = size_min

    def score(self, data_in):
        成功率合计 = 0
        范围 = []
        范围成功率 = []
        data = np.array(data_in)
        for a in range(self.__size_min, self.__size_max+1):
            for b in range(self.__size_min, self.__size_max+1):
                for c in range(self.__size_min, self.__size_max+1):
                    sum_part1 = self.__cal(data, 0, a)
                    sum_part2 = self.__cal(data, a, b)
                    sum_part3 = self.__cal(data, a + b, c)
                    sum_all1 = np.sum(data[0:a][:])
                    sum_all2 = np.sum(data[0:b][:])
                    sum_all3 = np.sum(data[0:c][:])
                    temp = sum_part1/sum_all1 + sum_part2/sum_all2 + sum_part3/sum_all3
                    if temp > 成功率合计:
                        成功率合计 = temp
                        范围 = [a, b, c]
                        范围成功率 = [sum_part1/sum_all1, sum_part2/sum_all2, sum_part3/sum_all3]
        return 范围, 范围成功率

    @staticmethod
    def __cal(data, start, size, mod='3*3'):
        res = 0
        if mod == '3*3':
            for i in range(size):
                for j in range(size):
                    res += data[start+i][start+j]
        elif mod == '3*1':
            for i in range(size):
                res += data[start+i][start+i]
        else:
            print('自动评价：没有 {} mod！！！'.format(mod))
            exit()
        return res


def excel(size_max=4, size_min=3):
    score = Score(size_max=size_max, size_min=size_min)
    wb = openpyxl.load_workbook(filename=path_excel)
    ws = wb.worksheets[0]
    excel_clear(ws)
    y = 2
    x = 4 * 0
    files = os.listdir(path)

    number = 0
    for file in files:
        if not os.path.isdir(path + '/' + file):
            number += 1
    bar = progressbar.ProgressBar(max_value=number*2, widgets=[
        '自动评价中: ',
        progressbar.Bar('>'),
        ' ',
        progressbar.Counter(format='%(value)02d/%(max_value)d'), ], )

    # for file in files:
    #     if not os.path.isdir(path + '/' + file):
    #         this = Get.get_data(path + '/' + file)
    #         res = score.score(this.result1)
    #         if file.split('_')[-3].find('告') == -1:
    #             ws.cell(row=y, column=1).value = file.split('_')[-3] + '_' + file.split('_')[-2]
    #         else:
    #             ws.cell(row=y, column=1).value = file.split('_')[-2]
    #         ws.cell(row=y, column=2+x).value = res[1][0]
    #         ws.cell(row=y, column=3+x).value = res[1][1]
    #         ws.cell(row=y, column=4+x).value = res[1][2]
    #         bar.update(y-2)
    #         y += 1

    y = 2
    x = 4 * 1
    for idd in range(1, 45):
        with DB() as db:
            sql = "SELECT name FROM good_id"
            res = db.read(sql)
            for name in res:
                result1 = get_data(name['name'])
                res = score.score(result1)
                if len(res[0]) == 0:
                    ws.cell(row=y, column=2 + x).value = 0
                    ws.cell(row=y, column=3 + x).value = 0
                    ws.cell(row=y, column=4 + x).value = 0
                else:
                    ws.cell(row=y, column=2 + x).value = res[1][0]
                    ws.cell(row=y, column=3 + x).value = res[1][1]
                    ws.cell(row=y, column=4 + x).value = res[1][2]
                # bar.update(y-2+number)
                y += 1
    try:
        wb.save(filename=path_excel)
    except PermissionError:
        print()
        print('保存 {} 文件时：缺少权限 or 文件已打开！！！'.format(path_excel))
        wb.close()
        exit()
    print('Score table has created!')
    # bar.finish()


def excel_clear(ws):
    y = 2
    for i in range(44):
        ws.cell(row=y + i, column=1).value = None
        ws.cell(row=y + i, column=2).value = None
        ws.cell(row=y + i, column=3).value = None
        ws.cell(row=y + i, column=4).value = None
        ws.cell(row=y + i, column=6).value = None
        ws.cell(row=y + i, column=7).value = None
        ws.cell(row=y + i, column=8).value = None


def get_data(name):
    result2 = []
    for i in range(30):
        data_temp = [list() for i in range(30)]
        for ii in range(30):
            data_temp[ii] = 0
        result2.append(deepcopy(data_temp))
    with DB() as db:
        sql = "SELECT id FROM data\
                    WHERE name = '{}' AND a = 0".format(name)
        res_id = db.read(sql)

        for j in range(len(res_id)):
            if (j - 1) < 0:
                head = 0
            else:
                head = res_id[j-1]['id']
            end = res_id[j]['id']
            sql = "SELECT pre FROM data\
                        WHERE id > {} AND id < {} AND name='{}'".format(head, end, name)
            res = db.read(sql)
            pre = [res[i]['pre'] for i in range(len(res))]
            for index, temp in enumerate(pre):
                if temp == 0:
                    continue
                result2[len(pre) - index - 1][temp-1] += 1
    return result2


if __name__ == '__main__':
    main()
