import openpyxl
import copy


class DrinkData:
    result1 = []
    result2 = []    # a+b-c
    data1 = []      # a
    data2 = []      # b
    data3 = []      # c
    data4 = []      # d
    package_mess = [list() for i in range(7)]   # 包装信息
    y = []
    逻辑库存 = []
    预计库存 = []

    def __init__(self):
        self.逻辑库存.clear()
        self.预计库存.clear()
        self.y.clear()
        self.result1.clear()
        self.result2.clear()
        self.data1.clear()
        self.data2.clear()
        self.data3.clear()
        self.data4.clear()
        self.package_mess.clear()
        self.package_mess = [list() for i in range(7)]


def main():
    path = "C:/Users/liang-jiashun/Desktop/001.xlsx"
    this = get_data(path, sheet=2)
    print(this)


def get_data(path, sheet=2, head_del=False):
    this = DrinkData()
    wb = openpyxl.load_workbook(filename=path, data_only=True)

    ws = wb.worksheets[0]
    this.package_mess[0].append(copy.deepcopy(ws.cell(row=17, column=2).value))
    材料 = ws.cell(row=17, column=8).value
    if 材料 == "铝合金":
        this.package_mess[1].append("アルミ")
    elif 材料 == "塑料":
        this.package_mess[1].append("ペットボトル")
    elif 材料 == "铝":
        this.package_mess[1].append("アルミ")
    elif 材料 == "马口铁":
        this.package_mess[1].append("ブリキ")
    else:
        this.package_mess[1].append("Other")
    this.package_mess[2].append(copy.deepcopy(ws.cell(row=17, column=11).value))
    this.package_mess[3].append(copy.deepcopy(ws.cell(row=17, column=13).value))
    this.package_mess[4].append(copy.deepcopy(ws.cell(row=17, column=15).value))
    this.package_mess[5].append(copy.deepcopy(ws.cell(row=17, column=17).value))
    wb1 = openpyxl.load_workbook(filename="C:\\Users\\Nao KeTeng\\OneDrive\\桌面\\Excel模板/测试商品硬度统计.xlsx", data_only=True)
    ws1 = wb1.worksheets[0]
    for i in range(len(this.package_mess[0])):
        this.package_mess[6].append("中")
    for i in range(100):
        if ws1.cell(row=2 + i, column=2).value == None:
            break
        if ws1.cell(row=2 + i, column=4).value == "偏软":
            for ii in range(len(this.package_mess[0])):
                if this.package_mess[0][ii] == ws1.cell(row=2 + i, column=2).value:
                    this.package_mess[6][ii] = "軟"
        elif ws1.cell(row=2 + i, column=4).value == "偏硬":
            for ii in range(len(this.package_mess[0])):
                if this.package_mess[0][ii] == ws1.cell(row=2 + i, column=2).value:
                    this.package_mess[6][ii] = "硬"
    wb1.close()

    for i in range(24):
        data_temp = [list() for i in range(24)]
        for ii in range(24):
            data_temp[ii] = 0
        this.result1.append(copy.deepcopy(data_temp))
    first = 0
    if sheet == 1:
        ws = wb.worksheets[1]
        x_start = 11
        y_start = 4
    elif sheet == 2 and len(wb.worksheets) > 2:
        ws = wb.worksheets[2]
        x_start = 1
        y_start = 2
        first = 1
    if len(wb.worksheets) == 2:
        ws = wb.worksheets[1]
        x_start = 11
        y_start = 4

    flag = 0
    for i in range(1000):
        if ws.cell(row=i + y_start, column=x_start).value == None:
            break
        if flag < 3:
            flag += 1
            continue
        if int(ws.cell(row=i + y_start, column=x_start).value) == 0:
            flag = 0
            first = 0
            continue
        if first == 1:
            continue
        aa = int(ws.cell(row=y_start + i, column=x_start).value)
        bb = int(ws.cell(row=y_start + i, column=x_start + 1).value)
        this.result1[aa - 1][bb - 1] += 1

    temp = []
    index = 1
    flag_last = False
    for i in range(1000):
        if ws.cell(row=i + y_start, column=x_start).value == None:
            break
        逻辑库存 = int(ws.cell(row=i + y_start, column=x_start).value)
        预计库存 = int(ws.cell(row=i + y_start, column=x_start + 1).value)
        try:
            aa = int(ws.cell(row=i + y_start, column=x_start + 2).value)
            bb = int(ws.cell(row=i + y_start, column=x_start + 1 + 2).value)
            cc = int(ws.cell(row=i + y_start, column=x_start + 2 + 2).value)
            dd = int(ws.cell(row=i + y_start, column=x_start + 3 + 2).value)
        except BaseException:
            pass
        if flag_last or (not head_del):
            try:
                this.data1.append(aa)
                this.data2.append(bb)
                this.data3.append(cc)
                this.data4.append(dd)
            except BaseException:
                pass
            this.逻辑库存.append(逻辑库存)
            this.预计库存.append(预计库存)
        flag_last = True
        if int(aa) == 0:
            flag_last = False
            this.result2.append(copy.deepcopy(temp))
            temp.clear()
            this.y.append(0)
            index = 1
            continue
        try:
            temp.append(aa + bb - cc)
        except BaseException:
            pass
        this.y.append(index)
        index += 1
    wb.close()
    this.y.reverse()
    return this


if __name__ == '__main__':
    main()
