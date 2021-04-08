from openpyxl import load_workbook
import shutil
import copy
import os
import progressbar
import csv
import time
from functools import wraps

PT試験仕様書目录 = 'C:/Users/liang-jiashun/Desktop/002'
EPTREE文件目录 = 'C:/Users/liang-jiashun/Desktop/003'
生成模板 = 'C:/Users/liang-jiashun/Desktop/PT一览表.xlsx'


class PTobject:
    def __init__(self, path, wb):
        self.機能 = []                            # PT式样书对应机能
        self.file = []                            # 文件名
        self.function = []                        # 函数名
        self.__機能_temp = self.__dealWith_機能Name(path)
        self.__file_temp = []
        self.wb = wb
        for i, ws in enumerate(wb.worksheets):
            if ws.title.find('試験一覧') != -1:
                self.start_sheet = i + 1          # PT式样书中 試験一覧 sheet
                break

    def get_mess(self):
        self.__get_file()

    def __get_file(self):                                       # 找到文件名
        for ws in self.wb.worksheets[self.start_sheet:]:
            if self.__get_start(ws) is not None:
                self.__file_temp = self.__dealWith_fileName(ws.title)
                self.__get_function(ws)

    def __get_function(self, ws):                               # 找到函数名
        if self.__get_start(ws) is None:
            return
        y, x = self.__get_start(ws)

        new = ' '
        while y < 200:                                          # PT式样书最大行数
            if ws.cell(row=y, column=x).value is None:
                y += 1
                continue
            else:
                old = new
                new = str(ws.cell(row=y, column=x).value)
                new = self.__dealWith_functionName(new)
            if new != old:
                self.function.append(copy.deepcopy(new))
                self.file.append(copy.deepcopy(self.__file_temp))
                self.機能.append(copy.deepcopy(self.__機能_temp))
            y += 1

    @staticmethod
    def __get_start(ws):
        for i in range(1, 30):
            for j in range(1, 5):
                temp = ws.cell(row=i, column=j).value
                if temp is None:
                    continue
                temp = str(temp)
                if temp.find('関数・メソッド') != -1:
                    return i+2, j
        else:
            return None

    @staticmethod
    def __dealWith_機能Name(name):
        if name.find('(') != -1:
            return name.split('(')[-1].split(')')[0]
        elif name.find('（')!= -1:
            return name.split('（')[-1].split('）')[0]

    @staticmethod
    def __dealWith_fileName(name):
        if name.find('PT') != -1:
            if name.find('(') != -1:
                return name.split('(')[-1].split(')')[0] + '.c'
            elif name.find('（') != -1:
                return name.split('（')[-1].split('）')[0] + '.c'
        else:
            return name + '.c'

    @staticmethod
    def __dealWith_functionName(name):
        return name.split('(')[0].split()[-1]


class EPTREEobject:
    def __init__(self, path, reader):
        # self.機能 = []
        self.file = []                            # 文件名
        self.function = []                        # 函数名
        self.reader = reader

    def get_mess(self):
        flag = False
        for temp in self.reader:
            if not flag:
                flag = True
                continue
            self.file.append(copy.deepcopy(temp[-1]))
            self.function.append(copy.deepcopy(temp[0]))


def writeExcel(pt, path):
    x = 1
    y = 2
    wb = load_workbook(filename=path)
    ws = wb.worksheets[0]
    for i in range(len(pt.function)):
        copy_set_cell(ws.cell(row=2, column=1), ws.cell(row=y, column=x), pt.機能[i])
        copy_set_cell(ws.cell(row=2, column=1), ws.cell(row=y, column=x + 1), pt.file[i])
        copy_set_cell(ws.cell(row=2, column=1), ws.cell(row=y, column=x + 2), pt.function[i])
        y += 1
    wb.save(filename=path)
    wb.close()


def copy_set_cell(cell_src, cell, value):
    cell.value = value
    if cell_src.has_style:
        cell.font = copy.copy(cell_src.font)
        cell.border = copy.copy(cell_src.border)
        cell.fill = copy.copy(cell_src.fill)
        cell.number_format = copy.copy(cell_src.number_format)
        cell.protection = copy.copy(cell_src.protection)
        cell.alignment = copy.copy(cell_src.alignment)


def 目录遍历(path, enable_createFile=True):
    def decorator(fun):
        def wrapper(*args):
            if not os.path.exists(path + '/test') and enable_createFile:  # 创建test目录
                os.makedirs(path + '/test')
            files = os.listdir(path)
            for i, file in enumerate(files):
                if os.path.isdir(path + '/' + file):
                    files.pop(i)

            if enable_createFile:
                mess_bar = 'PT表格导出中'
            else:
                mess_bar = '读取EPTREE表格中'
            bar = progressbar.ProgressBar(max_value=len(files), widgets=[
                mess_bar,
                progressbar.Bar('>'),
                ' ',
                progressbar.Counter(format='%(value)02d/%(max_value)d'), ], )

            for i, file in enumerate(files):
                bar.update(i+1)
                path_temp = ''
                if enable_createFile:
                    path_temp = path + '/test/PT一览表_' + file
                    shutil.copy(生成模板, path_temp)
                fun(path+ '/' + file, path_temp, args[0])
            bar.finish()
        return wrapper
    return decorator


@目录遍历(PT試験仕様書目录, enable_createFile=True)
def PT_test(*args):
    wb = load_workbook(filename=args[0], read_only=True)
    pt = PTobject(args[0], wb)
    pt.get_mess()
    wb.close()

    writeExcel(pt, args[1])

    pt_mess = args[2]
    pt_mess[0].extend(pt.機能)
    pt_mess[1].extend(pt.file)
    pt_mess[2].extend(pt.function)


@目录遍历(EPTREE文件目录, enable_createFile=False)
def EPTREE_test(*args):
    with open(args[0], 'rt') as fp:
        reader = csv.reader(fp)
        ep = EPTREEobject(args[0], reader)
        ep.get_mess()
    ep_mess = args[2]
    # ep_mess[0].extend(ep.機能)
    ep_mess[1].extend(ep.file)
    ep_mess[2].extend(ep.function)


def Test(pt_mess, ep_mess):
    path = 生成模板.split('/')
    path.pop(-1)
    path = '/'.join(path)
    path = '{}/PT一览表{}.xlsx'.format(path, time.strftime("%Y-%m-%d", time.localtime()))
    shutil.copy(生成模板, path)
    x = 1
    y = 2
    wb = load_workbook(filename=path)
    ws = wb.worksheets[0]
    for i in range(len(ep_mess[2])):
        copy_set_cell(ws.cell(row=2, column=1), ws.cell(row=y, column=x + 1), ep_mess[1][i])
        copy_set_cell(ws.cell(row=2, column=1), ws.cell(row=y, column=x + 2), ep_mess[2][i])
        temp = search_same(pt_mess, None, ep_mess[1][i], ep_mess[2][i])
        if temp:
            copy_set_cell(ws.cell(row=2, column=1), ws.cell(row=y, column=x + 3), '●')
            copy_set_cell(ws.cell(row=2, column=1), ws.cell(row=y, column=x), temp)
        else:
            copy_set_cell(ws.cell(row=2, column=1), ws.cell(row=y, column=x + 3), '×')
            copy_set_cell(ws.cell(row=2, column=1), ws.cell(row=y, column=x), '/')

        y += 1
    wb.save(filename=path)
    wb.close()


def search_same(src, a, b, c):

    temp = copy.deepcopy(src[2])
    for i in range(len(temp)):
        temp[i] = temp[i].lower()

    index = -1
    while index + 1 < len(temp):
        try:
            index = temp[index + 1:].index(c.lower()) + index + 1
        except ValueError:
            return None
        if src[1][index].lower() == b.lower():
            ret = copy.deepcopy(src[0][index])
            src[0].pop(index)
            src[1].pop(index)
            src[2].pop(index)
            return ret
    return None


def use_time(mode='s'):
    def decorator(fun):
        @wraps(fun)
        def wrapper(*args):
            start = time.time()
            res = fun(*args)
            if mode == 'ms':
                print("{name}:time is {time}ms".format(name=wrapper.__name__, time=(time.time()-start)*1000))
            else:
                print("{name}:time is {time}s".format(name=wrapper.__name__, time=time.time() - start))
            return res
        return wrapper
    return decorator


@use_time()
def main():
    pt_mess = [list() for i in range(3)]        # [機能, file, function]
    eptree_mess = [list() for i in range(3)]    # [機能, file, function]

    PT_test(pt_mess)                # PT
    EPTREE_test(eptree_mess)        # EPTREE
    Test(pt_mess, eptree_mess)      # PT 和 EPTREE 匹配结果


if __name__ == '__main__':
    main()
