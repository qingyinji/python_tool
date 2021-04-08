import numpy as np
import openpyxl
import os

path = "C:/Users/liang-jiashun/Desktop/库存感知测试结果汇总_最终"
way = "old"


def main():
    path_file = path
    if way == "new":
        path_file = path + "/" + "test"
    files = os.listdir(path_file)
    for file in files:
        if not os.path.isdir(path_file + "/" + file):
            print(file)
            result1, result2, name = get_data(path_file + "/" + file)
            std = np.std(result2, ddof=1)
            rmse = get_rmse(result1, result2)
            # print(name, rmse, std/rmse)
            # print(rmse)
            print(std/rmse)


def get_data(path):
    wb = openpyxl.load_workbook(filename=path, data_only=True)
    if len(wb.worksheets) == 2:
        ws = wb.worksheets[1]
        y_start = 4
        x_start = 11
    else:
        ws = wb.worksheets[2]
        y_start = 2
        x_start = 1
    result1 = []
    result2 = []
    for i in range(1000):
        if ws.cell(row=y_start+i, column=x_start).value == None:
            break
        aa = int(ws.cell(row=y_start + i, column=x_start).value)
        bb = int(ws.cell(row=y_start + i, column=x_start+1).value)
        if 0 < aa < 9 and bb < 50:
            result1.append(aa)
            result2.append(bb)
    return result1, result2, wb.worksheets[0].cell(row=17, column=2).value


def get_rmse(result1, result2):
    num_all = 0
    for i in range(len(result1)):
        num_all += (result1[i]-result2[i])**2
    return (num_all/len(result1))**0.5


if __name__ == '__main__':
    main()