import openpyxl
import os

path = "C:/Users/liang-jiashun/Desktop/0001"
files = os.listdir(path)
for file in files:
    if not os.path.isdir(path + "/" + file):
        print(file)
        wb = openpyxl.load_workbook(filename=path + "/" + file)
        ws1 = wb.worksheets[1]
        index = 0
        flag = 0
        for ii in range(1000):
            if ws1.cell(row= 4+index, column= 11).value == None:
                break
            if flag < 3:
                if int(ws1.cell(row=4 + index, column=11 + 1).value) != 512 and int(
                        ws1.cell(row=4 + index, column=11 + 1).value) != 1024:
                    ws1.cell(row=4 + index, column=11+1).value = 512
            flag += 1
            if int(ws1.cell(row=4 + index, column=11).value) == 0:
                flag = 0
            index += 1
        print(index)

        if len(wb.worksheets) > 2:
            ws2 = wb.worksheets[2]
            index = 0
            flag = 0
            for ii in range(1000):
                aa = ws2.cell(row=2 + index, column=1).value
                bb = ws2.cell(row=2 + index, column=2).value
                if ws2.cell(row=2 + index, column=1).value == None:
                    break
                if flag < 3:
                    if int(ws2.cell(row=2 + index, column=1 + 1).value) != 512 and int(
                            ws2.cell(row=2 + index, column=1 + 1).value) != 1024:
                        ws2.cell(row=2 + index, column=1 + 1).value = 512
                flag += 1
                if int(ws2.cell(row=2 + index, column=1).value) == 0:
                    flag = 0
                index += 1
            print(index)
        wb.save(filename=path + "/" + file)
        wb.close()