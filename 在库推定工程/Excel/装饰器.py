import time
from functools import wraps
import WriteExcel
import os
import openpyxl
import 自动评价 as Auto


def 列表循环(path, enable_createFile=True):
    def decorator(fun):
        def wrapper():
            files = os.listdir(path)
            for file in files:
                if os.path.isdir(path + '/' + file):
                    continue
                result = fun(path, '/' + file)

                if enable_createFile:
                    wb_temp = openpyxl.load_workbook(filename=path + "/test/test_" + file)
                    ws_temp = wb_temp.worksheets[1]
                    i = 0
                    for tempp in result:
                        for ii in range(3):
                            ws_temp.cell(row=4 + i, column=11).value = len(tempp) - ii + 3
                            ws_temp.cell(row=4 + i, column=12).value = 512
                            i += 1
                        for ii in range(len(tempp)):
                            ws_temp.cell(row=4 + i, column=11).value = len(tempp) - ii
                            ws_temp.cell(row=4 + i, column=12).value = tempp[ii]
                            i += 1
                        ws_temp.cell(row=4 + i, column=11).value = 0
                        ws_temp.cell(row=4 + i, column=12).value = 0
                        i += 1
                    wb_temp.save(filename=path + "/test/test_" + file)
        return wrapper
    return decorator


def 自动评价(path, path_excel, size_min=3, size_max=4, enable=True):
    def decorator(fun):
        @wraps(fun)
        def wrapper(*args):
            fun(*args)
            if enable:
                Auto.excel(path, path_excel, size_min=size_min, size_max=size_max)
        return wrapper
    return decorator


def use_time(mode='s'):
    def decorator(fun):
        @wraps(fun)
        def wrapper(*args):
            start = time.time()
            res = fun(*args)
            if mode == 'ms':
                print("{name}:time is {time}ms".format(name=wrapper.__name__, time=(time.time()-start)*1000))
            else:
                print("{name}:time is {time}s".format(name=wrapper.__name__, time=time.time() - start))
            return res
        return wrapper
    return decorator


def cache_data(fun):
    cache = []

    @wraps(fun)
    def wrapper(*args):
        temp = fun(*args)
        if temp[0] == -1:
            return cache
        if temp[0] == 0:
            cache.clear()
            return None
        cache.append(temp)
        return cache
    return wrapper


def output(fun):
    def wrapper(*args):
        path, result = fun(*args)
        WriteExcel.writeexcel(result, path)
    return wrapper
