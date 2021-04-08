import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Border,Side
from openpyxl.styles import Alignment
import copy


def writeexcel(result, path):
    x_start = 25
    y_start = 25
    result.reverse()
    wb = openpyxl.load_workbook(filename=path)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])

    for i in range(len(result)):
        if ws.cell(row=y_start-i, column=x_start-result[i]+1).value == None:
            ws.cell(row=y_start - i, column=x_start - result[i] + 1).value = 0
        ws.cell(row=y_start-i, column=x_start-result[i]+1).value += 1
    result.reverse()
    wb.save(filename=path)
    wb.close()


def initexcel(path):
    x_start = 25
    y_start = 25
    wb = openpyxl.load_workbook(filename=path)
    ws = wb.worksheets[0]

    for i in range(2, x_start + 1):
        for j in range(2, y_start + 1):
            ws.cell(row=j, column=i).value = 0

    for i in range(300):
        wb.worksheets[1].cell(row=4+i, column=11).value = 0
        wb.worksheets[1].cell(row=4 + i, column=12).value = 0

    wb.save(filename=path)
    wb.close()


def set_style(work_sheet, cell, data):
    low = "EEB4B4"
    mid = "EEE685"
    high = "C1FFC1"
    other = "FFFFFF"
    if 0 < data < 0.5:
        color = low
    elif 0.5 <= data < 0.8:
        color = mid
    elif 0.8 <= data <= 1:
        color = high
    else:
        color = other
    cell.style = work_sheet.cell(row=2, column=3).style
    cell.border = Border(left=Side(border_style='thin', color ='000000'), right = Side(
        border_style='thin', color ='000000'), top = Side(border_style='thin', color ='000000'), bottom = Side(
        border_style='thin', color ='000000'))
    cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=False)
    cell.fill = PatternFill(fgColor=color, fill_type="solid")


def set_cell(work_sheet, y, x, value, *args):
    work_sheet.cell(row=y, column=x).value = value
    if args:
        set_style(work_sheet, work_sheet.cell(row=y, column=x), value)
    else:
        work_sheet.cell(row=y, column=x).alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        work_sheet.cell(row=y, column=x).border = Border(left=Side(border_style='thin', color='000000'), right=Side(
            border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(
            border_style='thin', color='000000'))


def set_cell1(work_sheet, y, x, value, *args):
    work_sheet.cell(row=y, column=x).value = value
    work_sheet.cell(row=y, column=x).alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    work_sheet.cell(row=y, column=x).border = Border(left=Side(border_style='thin', color='000000'), right=Side(
            border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(
            border_style='thin', color='000000'))


def copy_set_cell(cell_src, cell, value):
    cell.value = value  # 赋值到ws2单元格
    if cell_src.has_style:  # 拷贝格式
        cell.font = copy.copy(cell_src.font)
        cell.border = copy.copy(cell_src.border)
        cell.fill = copy.copy(cell_src.fill)
        cell.number_format = copy.copy(cell_src.number_format)
        cell.protection = copy.copy(cell_src.protection)
        cell.alignment = copy.copy(cell_src.alignment)


def copy_sheet(wb_src,ws_src, name):
    ws_book_new = wb_src.create_sheet(name)
    ws_book = ws_src
    max_row = ws_book.max_row  # 最大行数
    max_column = ws_book.max_column  # 最大列数

    wm = zip(ws_book.merged_cells)  # 开始处理合并单元格
    if len(list(wm)) > 0:
        for i in range(0, len(list(wm))):
            cell2 = str(wm[i]).replace('(<MergeCell ', '').replace('>,)', '')
            print("MergeCell : %s" % cell2)
            ws_book_new.merge_cells(cell2)

    for m in range(1, max_row + 1):
        ws_book_new.row_dimensions[m].height = ws_book.row_dimensions[m].height
        for n in range(1, 1 + max_column):
            if n < 27:
                c = chr(n + 64).upper()  # ASCII字符,chr(65)='A'
            else:
                if n < 677:
                    c = chr(divmod(n, 26)[0] + 64) + chr(divmod(n, 26)[1] + 64)
                else:
                    c = chr(divmod(n, 676)[0] + 64) + chr(divmod(divmod(n, 676)[1], 26)[0] + 64) + chr(
                        divmod(divmod(n, 676)[1], 26)[1] + 64)
            i = '%s%d' % (c, m)  # 单元格编号
            if m == 1:
                ws_book_new.column_dimensions[c].width = ws_book.column_dimensions[c].width
            try:
                getattr(ws_book.cell(row=m, column=n), "value")
                cell1 = ws_book[i]  # 获取data单元格数据
                ws_book_new[i].value = cell1.value  # 赋值到ws2单元格
                if cell1.has_style:  # 拷贝格式
                    ws_book_new[i].font = copy.copy(cell1.font)
                    ws_book_new[i].border = copy.copy(cell1.border)
                    ws_book_new[i].fill = copy.copy(cell1.fill)
                    ws_book_new[i].number_format = copy.copy(cell1.number_format)
                    ws_book_new[i].protection = copy.copy(cell1.protection)
                    ws_book_new[i].alignment = copy.copy(cell1.alignment)
            except AttributeError as e:
                print("cell(%s) is %s" % (i, e))
                continue
