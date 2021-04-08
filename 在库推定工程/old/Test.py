import os
import shutil
import copy
import openpyxl

import QuantityEstimate

path_src = "C:/Users/liang-jiashun/Desktop/库存感知测试结果汇总_最终"
path_sample = "C:/Users/liang-jiashun/Desktop/Excel模板"

path_sample_test = path_sample + "/" + "test.xlsx"
path_src_test = path_src + "/test"
path_src_test_test = path_src_test + "/" + "test.xlsx"

data = []
flag = 1
index = 0
index_y = 0
num = 0
num_group = 0
num_group_s = 0
wb_test = []

if not os.path.exists(path_src_test):             # 创建test目录
    os.makedirs(path_src_test)

files = os.listdir(path_src)           # 文件遍历
for file in files:
    if not os.path.isdir(path_src + "/" + file):
        print(file)
        if os.path.exists(path_src_test + "/test_" + file):
            os.remove(path_src_test + "/test_" + file)
        shutil.copy(path_sample_test, path_src_test)
        os.rename(path_src_test_test, path_src_test + "/test_" + file)
        wb = openpyxl.load_workbook(filename=path_src + "/" + file, data_only=True)

        if len(wb.worksheets) == 2:
            x_start = 13
            y_start = 4
            ws = wb.worksheets[1]
            for _ in range(1000):
                if ws.cell(row=_+y_start, column=x_start).value == None:
                    break
                elif int(ws.cell(row=_+y_start, column=x_start).value) == 0:
                    num_group += 1
                num += 1
            num_group_s = int(num / num_group)
        elif len(wb.worksheets) == 3:
            ws = wb.worksheets[2]
            for _ in range(1000):
                if ws.cell(row=_+2, column=1).value == None:
                    break
                elif int(ws.cell(row=_+2, column=1).value) == 0:
                    num_group += 1
                num += 1
            num_group_s = int(num / num_group)
            x_start = 3
            y_start = 2

        temp = []
        for i in range(num):
            aa = int(ws.cell(row=y_start + i, column=x_start).value)
            bb = int(ws.cell(row=y_start + i, column=x_start + 1).value)
            cc = int(ws.cell(row=y_start + i, column=x_start + 2).value)
            if aa != 0 and bb != 0 and cc != 0:
                temp.append(aa + bb - cc)
            else:
                data.append(copy.deepcopy(temp))
                temp.clear()
        data_temp = []
        data_temp.append(copy.deepcopy(data))
        data.pop(0)

        result = QuantityEstimate.quantityestimate(data_temp, data, path_src_test + "/test_" + file)
        print(result)

        data.clear()
        data_temp.clear()
        wb.close()
        num = 0
        num_group = 0
        num_group_s = 0

        wb_temp = openpyxl.load_workbook(filename=path_src_test + "/test_" + file)
        ws_temp = wb_temp.worksheets[1]
        i = 0
        for tempp in result:
            for ii in range(3):
                ws_temp.cell(row=4 + i, column=11).value = len(tempp) - ii + 3
                ws_temp.cell(row=4 + i, column=12).value = 512
                i += 1
            for ii in range(len(tempp)):
                ws_temp.cell(row=4+i, column=11).value = len(tempp) - ii
                ws_temp.cell(row=4 + i, column=12).value = tempp[ii]
                i += 1
            ws_temp.cell(row=4 + i, column=11).value = 0
            ws_temp.cell(row=4 + i, column=12).value = 0
            i += 1
        wb_temp.save(filename=path_src_test + "/test_" + file)
