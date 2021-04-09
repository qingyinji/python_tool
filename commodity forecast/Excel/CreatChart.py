import os
import openpyxl
import copy
import shutil
from openpyxl.drawing.image import Image
import zipfile
import matplotlib.pyplot as plt
import numpy as np

path_bath = "库存感知测试结果汇总_最终"
path_src = "C:/Users/liang-jiashun/Desktop" + "/" + path_bath
path_sample = "C:/Users/liang-jiashun/Desktop/Excel模板"
path_sample_chart = path_sample + "/" + "chart.xlsx"
stdtemp = []


def main():
    files = os.listdir(path_src)
    for file in files:
        if not os.path.isdir(path_src + "/" + file):
            # print(file)
            result, data1, data2, data3, name = get_data(path_src + "/" + file)
            # write_excel(result, file, data1, data2, data3, name)
            # print(result)
    print(stdtemp)
    plt.figure()
    for i in stdtemp:
        plt.scatter(range(3), i)
    plt.show()
    # add_image(path_src, path_src + "/test/" + "image")


def write_excel(result, file, data1, data2, data3, name):
    path_chart = path_src + "/test/image/" + "chart_" + file
    shutil.copy(path_sample_chart, path_chart)
    wb = openpyxl.load_workbook(filename=path_chart)
    ws = wb.worksheets[1]

    wb.worksheets[0].cell(row=2, column=2).value = name
    for i in range(len(data1)):
        wb.worksheets[2].cell(row=2+i, column=1).value = data1[i]
        wb.worksheets[2].cell(row=2 + i, column=2).value = data2[i]
        wb.worksheets[2].cell(row=2 + i, column=3).value = data3[i]

    for i in range(len(result[0])):
        ws.cell(row=2+i, column=1).value = len(result[0]) - i
    for i, temp in enumerate(result):
        for ii, data in enumerate(temp):
            ws.cell(row=2 + ii, column=2+i).value = data
    wb.save(filename=path_chart)


def get_data(path):
    wb = openpyxl.load_workbook(filename=path, data_only=True)
    ws = wb.worksheets[1]
    y_start = 4
    x_start = 13
    result = []
    data1 = []
    data2 = []
    data3 = []
    temp = []
    for i in range(1000):
        if ws.cell(row=i + y_start, column=x_start).value == None:
            break
        aa = int(ws.cell(row=i + y_start, column=x_start).value)
        bb = int(ws.cell(row=i + y_start, column=x_start + 1).value)
        cc = int(ws.cell(row=i + y_start, column=x_start + 2).value)
        data1.append(aa)
        data2.append(bb)
        data3.append(cc)
        if int(aa) == 0:
            result.append(copy.deepcopy(temp))
            temp.clear()
            continue
        temp.append(aa+bb-cc)
    result = data_analysis(result, data1, data2, data3, 2)
    return result, data1, data2, data3, wb.worksheets[0].cell(row=17, column=2).value


def data_analysis(result, data1, data2, data3, flag):
    aa = [0, 0, 0]
    aa[0] = np.std(result[0][0:5])
    aa[1] = np.std(result[0][6:10])
    aa[2] = np.std(result[0][10:])
    for i in range(len(result)):
        aa[0] = (np.std(result[i][0:5])+aa[0])/2
        aa[1] = (np.std(result[i][6:10])+aa[1])/2
        aa[2] = (np.std(result[i][10:])+aa[2])/2
    stdtemp.append(copy.deepcopy(aa))
    print(aa)
    # temp = np.array(result).T
    # print(aa)
    # plt.figure()
    # plt.scatter(range(3), aa)
    # plt.show()
    # exit()


def add_image(path_src, path_file):
    name_file = os.listdir(path_file)

    path_image = "C:/Users/liang-jiashun/Desktop/Excel模板/image"
    shutil.rmtree(path_image)
    os.makedirs(path_image)
    files = os.listdir(path_src)  # 文件遍历
    for file in files:
        if file != "test":
            shutil.copy(path_src+"/"+file, path_image)

    index = 0
    files = os.listdir(path_image)           # 文件遍历
    for file in files:
        if not os.path.isdir(path_image + "/" + file):
            wb = openpyxl.load_workbook(filename=path_file + "/" +name_file[index])
            ws = wb.worksheets[0]

            file_name = os.path.basename(path_image + "/" + file)  # 获取文件名
            new_name = str(file_name.split('.')[0]) + ".zip"  # 新的文件名，命名为：xxx.zip
            dir_path = os.path.dirname(path_image + "/" + file)  # 获取文件所在目录
            new_path = dir_path+"/"+new_name  # 新的文件路径
            os.rename(path_image + "/" + file, new_path)

            file_zip = zipfile.ZipFile(new_path, 'r')
            zipdir = path_image+"/"+str(file.split('.')[0])  # 获取文件所在目录
            for files in file_zip.namelist():
                file_zip.extract(files, zipdir)  # 解压到指定文件目录
            file_zip.close()
            os.remove(new_path)

            pic_dir = 'xl'+"/"+'media'  # excel变成压缩包后，再解压，图片在media目录
            pic_path = path_image+"/"+str(file.split('.')[0])+"/"+pic_dir
            image = pic_path + "/" +"image1.jpeg"

            print(image)
            img1 = Image(image)
            img1.width, img1.height = 73*2, 73*2
            ws.add_image(img1,'A5')
            wb.save(filename=path_file + "/" +name_file[index])
            index += 1
    wb.close()


if __name__ == '__main__':
    main()