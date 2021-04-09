import os
import openpyxl
import copy
import shutil
from openpyxl.drawing.image import Image
import zipfile
from openpyxl.chart import LineChart, Reference

path_bath = "库存感知测试结果汇总_最终"
path_src = "C:/Users/liang-jiashun/Desktop" + "/" + path_bath
path_sample = "C:/Users/liang-jiashun/Desktop/Excel模板"

wb = openpyxl.load_workbook(filename=path_src + "/test/image/" + "0011.xlsx")


def main():
    index = 0
    files = os.listdir(path_src)
    for file in files:
        if not os.path.isdir(path_src + "/" + file):
            print(file)
            result, data1, data2, data3, name = get_data(path_src + "/" + file)
            write_excel(result, file, data1, data2, data3, name, index)
            index += 1
    wb.save(filename=path_src + "/test/image/" + "0011.xlsx")
    add_image(path_src, path_src + "/test/" + "image")


def chart(ws):
    data = Reference(ws, min_col=2, max_col=11, min_row=2, max_row=23)  # 引用数据系列，用作折线图表的数据值。
    c = LineChart()
    c.add_data(data, titles_from_data=True)
    dates = Reference(ws, min_col=1, min_row=2, max_row=23)
    c.set_categories(dates)
    c.height = 12
    c.width = 22
    ws.add_chart(c, 'A29')


def write_excel(result, file, data1, data2, data3, name, index):
    ws = wb.worksheets[index]
    if 21-len(result[0])>0:
        for i in range(21):
            for ii in range(21-len(result[0])):
                ws.cell(row=i+3,column=16+ii).value='-'
    ws.title = name
    chart(ws)
    for i in range(len(result)):
        for ii in range(21):
            ws.cell(row=3 + ii, column=2+i).value = None
    for i, temp in enumerate(result):
        for ii, data in enumerate(temp):
            ws.cell(row=24 - len(result[0]) + ii, column=2+i).value = data


def get_data(path):
    wb = openpyxl.load_workbook(filename=path, data_only=True)
    ws = wb.worksheets[1]
    y_start = 4
    x_start = 13
    result = []
    data1 = []
    data2 = []
    data3 = []
    temp = []
    for i in range(1000):
        if ws.cell(row=i + y_start, column=x_start).value == None:
            break
        aa = int(ws.cell(row=i + y_start, column=x_start).value)
        bb = int(ws.cell(row=i + y_start, column=x_start + 1).value)
        cc = int(ws.cell(row=i + y_start, column=x_start + 2).value)
        data1.append(aa)
        data2.append(bb)
        data3.append(cc)
        if int(aa) == 0:
            result.append(copy.deepcopy(temp))
            temp.clear()
            continue
        temp.append(aa+bb-cc)
    return result, data1, data2, data3, wb.worksheets[0].cell(row=17, column=2).value


def add_image(path_src, path_file):
    path_image = "C:/Users/liang-jiashun/Desktop/Excel模板/image"
    shutil.rmtree(path_image)
    os.makedirs(path_image)
    files = os.listdir(path_src)  # 文件遍历
    for file in files:
        if file != "test":
            shutil.copy(path_src+"/"+file, path_image)

    index = 0
    files = os.listdir(path_image)           # 文件遍历
    for file in files:
        if not os.path.isdir(path_image + "/" + file):
            ws = wb.worksheets[index]

            file_name = os.path.basename(path_image + "/" + file)  # 获取文件名
            new_name = str(file_name.split('.')[0]) + ".zip"  # 新的文件名，命名为：xxx.zip
            dir_path = os.path.dirname(path_image + "/" + file)  # 获取文件所在目录
            new_path = dir_path+"/"+new_name  # 新的文件路径
            os.rename(path_image + "/" + file, new_path)

            file_zip = zipfile.ZipFile(new_path, 'r')
            zipdir = path_image+"/"+str(file.split('.')[0])  # 获取文件所在目录
            for files in file_zip.namelist():
                file_zip.extract(files, zipdir)  # 解压到指定文件目录
            file_zip.close()
            os.remove(new_path)

            pic_dir = 'xl'+"/"+'media'  # excel变成压缩包后，再解压，图片在media目录
            pic_path = path_image+"/"+str(file.split('.')[0])+"/"+pic_dir
            image = pic_path + "/" +"image1.jpeg"

            print(image)
            img1 = Image(image)
            img1.width, img1.height = 73*2, 73*2
            ws.add_image(img1,'P33')
            wb.save(path_src + "/test/image/" + "0011.xlsx")
            index += 1
    wb.close()


if __name__ == '__main__':
    main()