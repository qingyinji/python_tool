import os
import shutil
import copy
import openpyxl
import WriteExcel
import AddImag

way = "new"
path_bath = "库存感知测试结果汇总_最终"
path_src = "C:/Users/liang-jiashun/Desktop" + "/" + path_bath
path_sample = "C:/Users/liang-jiashun/Desktop/Excel模板"
path_sample_book = path_sample + "/" + "報告書自動生成テンプレート_JP1.xlsx"


def main():
    result = []
    if way == "old":
        path = path_src
    else:
        path = path_src + "/" + "test"
    files = os.listdir(path)
    for file in files:
        if not os.path.isdir(path + "/" + file):
            print(file)
            data = get_data(path + "/" + file)
            result.append(copy.deepcopy(data_analysis(data, 1)))
    mess = get_mess()
    path = write_excel(result, mess, 1)
    AddImag.add_image(path_src, path)


# --------------------------------------------------------------------------------------------
def get_data(path):
    group = 24
    wb = openpyxl.load_workbook(filename=path, data_only=True)
    data = []
    for i in range(group):
        data_temp = [list() for i in range(group)]
        for ii in range(24):
            data_temp[ii] = 0
        data.append(copy.deepcopy(data_temp))

    ws = wb.worksheets[1]
    x_start = 11
    y_start = 4
    flag = 0
    for i in range(1000):
        if ws.cell(row=i + y_start, column=x_start).value == None:
            break
        if flag < 3:
            flag += 1
            continue
        if int(ws.cell(row=i + y_start, column=x_start).value) == 0:
            flag = 0
            continue
        aa = int(ws.cell(row=y_start + i, column=x_start).value)
        bb = int(ws.cell(row=y_start + i, column=x_start + 1).value)
        data[aa - 1][bb - 1] += 1
    wb.close()
    return data


# --------------------------------------------------------------------------------------------
def get_mess():
    package_mess = [list() for i in range(7)]  # 包装信息
    files = os.listdir(path_src)  # 文件遍历
    for file in files:
        if not os.path.isdir(path_src + "/" + file):
            wb_temp = openpyxl.load_workbook(filename=path_src + "/" + file, data_only=True)
            ws_temp = wb_temp.worksheets[0]
            package_mess[0].append(copy.deepcopy(ws_temp.cell(row=17, column=2).value))
            材料 = ws_temp.cell(row=17, column=8).value
            if 材料 == "铝合金":
                package_mess[1].append("合金アルミ")
            elif 材料 == "塑料":
                package_mess[1].append("ペットボトル")
            elif 材料 == "铝":
                package_mess[1].append("アルミ")
            elif 材料 == "马口铁":
                package_mess[1].append("ブリキ")
            else:
                package_mess[1].append("Other")
            package_mess[2].append(copy.deepcopy(ws_temp.cell(row=17, column=11).value))
            package_mess[3].append(copy.deepcopy(ws_temp.cell(row=17, column=13).value))
            package_mess[4].append(copy.deepcopy(ws_temp.cell(row=17, column=15).value))
            package_mess[5].append(copy.deepcopy(ws_temp.cell(row=17, column=17).value))
            wb_temp.close()

    wb = openpyxl.load_workbook(filename=path_sample + "/" + "测试商品硬度统计.xlsx", data_only=True)
    ws = wb.worksheets[0]
    for i in range(len(package_mess[0])):
        package_mess[6].append("中")
    for i in range(100):
        if ws.cell(row=2+i, column=2).value == None:
            break
        if ws.cell(row=2+i, column=4).value == "偏软":
            for ii in range(len(package_mess[0])):
                if package_mess[0][ii] == ws.cell(row=2+i, column=2).value:
                    package_mess[6][ii] = "軟"
        elif ws.cell(row=2+i, column=4).value == "偏硬":
            for ii in range(len(package_mess[0])):
                if package_mess[0][ii] == ws.cell(row=2+i, column=2).value:
                    package_mess[6][ii] = "硬"
    wb.close()
    return package_mess


# --------------------------------------------------------------------------------------------
def data_analysis(data, flag):
    result = [[], []]
    if flag == 1:
        for i in range(3):
            result[0].append(float((data[i*3][i*3]+data[i*3+1][i*3+1]+data[i*3+2][i*3+2])/30))
            result[1].append(float((sum(data[i*3][i*3:i*3+3])+sum(data[i*3+1][i*3:i*3+3])+sum(data[i*3+2][i*3:i*3+3]))/30))
        return result
    elif flag == 2:
        for i in range(3):
            result[0].append(float((data[i * 5][i * 5] + data[i * 5 + 1][i * 5 + 1] + data[i * 5 + 2][i * 5 + 2] +
                                    data[i * 5 + 3][i * 5 + 3] + data[i * 5 + 4][i * 5 + 4]) / 50))
            result[1].append(float((sum(data[i * 5][i * 5:i * 5 + 5]) + sum(data[i * 5 + 1][i * 5:i * 5 + 5]) + sum(
                data[i * 5 + 2][i * 5:i * 5 + 5]) + sum(data[i * 5 + 3][i * 5:i * 5 + 5]) + sum(
                data[i * 5 + 4][i * 5:i * 5 + 5])) / 50))
        return result


# --------------------------------------------------------------------------------------------
def write_excel(result, mess, flag):
    path_book = path_sample + "/" + "報告書_" + path_bath + "1.xlsx"
    if os.path.exists(path_book):
        os.remove(path_book)
    shutil.copy(path_sample_book, path_book)
    wb_book = openpyxl.load_workbook(filename=path_book)

    if flag == 1:
        for i in range(len(result)):
            cell_src0 = wb_book.worksheets[0].cell(row=4, column=2)
            cell_src1 = wb_book.worksheets[0].cell(row=4, column=2)
            cell0 = wb_book.worksheets[0].cell(row=4 + i, column=2)
            cell1 = wb_book.worksheets[1].cell(row=4 + i, column=2)
            WriteExcel.copy_set_cell(cell_src0, cell0, mess[0][i])
            WriteExcel.copy_set_cell(cell_src1, cell1, mess[0][i])

            for ii in range(13):
                cell_src = wb_book.worksheets[1].cell(row=4, column=3 + ii)
                cell = wb_book.worksheets[1].cell(row=4 + i, column=3 + ii)
                if ii < 6:
                    WriteExcel.copy_set_cell(cell_src, cell, mess[ii+1][i])
                else:
                    WriteExcel.copy_set_cell(cell_src, cell, '')
            temp = str(4 + i)
            cell = wb_book.worksheets[1].cell(row=4 + i, column=9)
            cell.value = "=IF(トータル情報!C" + temp + "=\"-\",\"-\",IF(トータル情報!C" + temp + ">0.8,\"高\",IF(トータル情報!C" + temp + ">0.5,\"中\",\"低\")))"
            cell = wb_book.worksheets[1].cell(row=4  + i, column=10)
            cell.value = "=IF(トータル情報!D" + temp + "=\"-\",\"-\",IF(トータル情報!D" + temp + ">0.8,\"高\",IF(トータル情報!D" + temp + ">0.5,\"中\",\"低\")))"
            cell = wb_book.worksheets[1].cell(row=4  + i, column=11)
            cell.value = "=IF(トータル情報!E" + temp + "=\"-\",\"-\",IF(トータル情報!E" + temp + ">0.8,\"高\",IF(トータル情報!E" + temp + ">0.5,\"中\",\"低\")))"
            cell = wb_book.worksheets[1].cell(row=4 + i, column=12)
            cell.value = "=IF(トータル情報!F" + temp + "=\"-\",\"-\",IF(トータル情報!F" + temp + ">0.8,\"高\",IF(トータル情報!F" + temp + ">0.5,\"中\",\"低\")))"
            cell = wb_book.worksheets[1].cell(row=4 + i, column=13)
            cell.value = "=IF(トータル情報!G" + temp + "=\"-\",\"-\",IF(トータル情報!G" + temp + ">0.8,\"高\",IF(トータル情報!G" + temp + ">0.5,\"中\",\"低\")))"
            cell = wb_book.worksheets[1].cell(row=4 + i, column=14)
            cell.value = "=IF(トータル情報!H" + temp + "=\"-\",\"-\",IF(トータル情報!H" + temp + ">0.8,\"高\",IF(トータル情報!H" + temp + ">0.5,\"中\",\"低\")))"
            # cell = wb_book.worksheets[1].cell(row=4  + i, column=14)
            # cell.value = "=IF(AND(COUNTIF(H" + temp + ":M" + temp + ",\"中\")=0,COUNTIF(H" + temp + ":M" + temp + ",\"低\")=0),\"〇\",IF(COUNTIF(H" + temp + ":M" + temp + ",\"低\")>0,\"×\",\"△\"))"
            for ii in range(3):
                cell = wb_book.worksheets[1].cell(row=4 + i, column=15+ii)
                a1 = result[i][0][ii]
                a2 = result[i][1][ii]
                # if a1 == 0 or a2 == 0:
                #     continue
                if a1 > 0.8 and a2 > 0.8:
                    WriteExcel.copy_set_cell(wb_book.worksheets[1].cell(row=4, column=14), cell, "高")
                elif (a1 > 0.8 and a2 > 0.5) or (a1 > 0.5 and a2 > 0.8):
                    WriteExcel.copy_set_cell(wb_book.worksheets[1].cell(row=4, column=14), cell, "中上")
                elif (a1 > 0.8 and a2 <= 0.5) or (a1 <= 0.5 and a2 > 0.8):
                    WriteExcel.copy_set_cell(wb_book.worksheets[1].cell(row=4, column=14), cell, "中")
                elif a1 > 0.5 and a2 > 0.5:
                    WriteExcel.copy_set_cell(wb_book.worksheets[1].cell(row=4, column=14), cell, "中")
                elif a1 > 0.5 or a2 > 0.5:
                    WriteExcel.copy_set_cell(wb_book.worksheets[1].cell(row=4, column=14), cell, "中下")
                else:
                    WriteExcel.copy_set_cell(wb_book.worksheets[1].cell(row=4, column=14), cell, "低")

            for ii in range(3):
                # if result[i][0][ii] != 0:
                #     temp = result[i][0][ii]
                # else:
                #     temp = '-'
                temp = result[i][0][ii]
                cell_src = wb_book.worksheets[0].cell(row=4, column=3 + ii)
                cell = wb_book.worksheets[0].cell(row=4 + i, column=3 + ii)
                WriteExcel.copy_set_cell(cell_src, cell, temp)
                # if result[i][1][ii] != 0:
                #     temp = result[i][1][ii]
                # else:
                #     temp = '-'
                temp = result[i][1][ii]
                cell_src = wb_book.worksheets[0].cell(row=4, column=6 + ii)
                cell = wb_book.worksheets[0].cell(row=4 + i, column=6 + ii)
                WriteExcel.copy_set_cell(cell_src, cell, temp)
        wb_book.save(filename=path_book)
        wb_book.close()
        return path_book


main()
