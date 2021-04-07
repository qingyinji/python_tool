import openpyxl

path = 'C:/Users/liang-jiashun/Desktop/test.xlsx'


def writeexcel(result):
    x_start = 16
    y_start = 14
    result.reverse()
    wb = openpyxl.load_workbook(filename=path)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])

    for i in range(len(result)):
        if ws.cell(row=y_start-i, column=x_start-result[i]+1).value == None:
            ws.cell(row=y_start - i, column=x_start - result[i] + 1).value = 0
        ws.cell(row=y_start-i, column=x_start-result[i]+1).value += 1
    wb.save(filename=path)
    result.reverse()


def initexcel():
    x_start = 16
    y_start = 14
    wb = openpyxl.load_workbook(filename=path)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])

    for i in range(2, x_start + 1):
        for j in range(2, y_start + 1):
            ws.cell(row=j, column=i).value = 0
    wb.save(filename=path)
